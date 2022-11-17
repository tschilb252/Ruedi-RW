import numpy as np
import pandas as pd
import os
from os import walk
from openpyxl import load_workbook
from datetime import datetime


def get_file_names(mydir, str=''):
    # gets a list of the xml filenames in the directory for a particular station
    f = []
    f2 = []
    for root, dirs, files in walk(mydir, topdown=True):
        for x in files:
            if x.endswith(str + '.xlsx') or x.endswith(str + '.xls') or x.endswith(str + '.xlsm'):
                f.append(os.path.join(root, x))
    return f


if __name__ == '__main__':
    inpath = 'C:\\Users\\JLanini\\Documents\\GitHub\\Ruedi-RW\\Data\\Accounting Spreadsheets\\'
    outpath = 'C:\\Users\\JLanini\\Documents\\GitHub\\Ruedi-RW\\Data\\'
    sheetnames = ['Contract Usage', 'Computations']

    dfcontracts = pd.DataFrame()
    dfall= pd.DataFrame()
    for f in get_file_names(inpath):
        for sheet in sheetnames:

            print('Processing ' + f)
            if sheet == 'Contract Usage':
                skip = list(range(5, 46))
                df = pd.read_excel(f, sheet_name=sheet, index_col=0, header=4, skiprows=skip, parse_dates=True)
                df.drop('Total Releases for Contractors', axis=1, inplace=True)
                dfout = dfcontracts.append(df)
                df=pd.DataFrame()
            else:
                skip = [0,1,2,8,9]
                df = pd.read_excel(f, sheet_name=sheet, index_col=1, skiprows=skip,
                                   parse_dates=True)
                #df.drop('Total Releases for Contractors', axis=1, inplace=True)
                dfcols=df.iloc[:4].replace(np.nan, '')
                column_names = [' '.join(col_items) for col_items in dfcols.iloc[:4].values.T]
                df = df.iloc[4:]
                df.columns = column_names
                df = df.loc[:, ~df.columns.duplicated()].copy()
                dfall = pd.concat([dfall,df])
                df=pd.DataFrame()
    dfall = dfall.sort_index()
    dfall=dfall.drop_duplicates(keep='last')
    dfall=dfall[dfall.index.duplicated() == False]
    new_date_range = pd.date_range(start="2005-04-01", end="2022-10-31", freq="D")
    dfall.reindex(new_date_range)
    dfcontracts = dfcontracts.sort_index()
    dfall.to_csv(outpath + 'AccountingDataTSAll.csv')
    dfcontracts.to_csv(outpath + 'AccountingDataTSContracts.csv')
